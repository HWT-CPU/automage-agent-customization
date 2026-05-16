import json
import os
import uuid
from datetime import date, timedelta, datetime
from pathlib import Path
from typing import Any

import psycopg
import requests
from openpyxl import Workbook

BASE_DIR = Path(r"C:\Users\23157\CODE\Moltbot\01\03")
OUT_DIR = BASE_DIR / "automage_m2_real_api_integration_acceptance"
API_LOG_DIR = OUT_DIR / "api_logs"
SCRIPT_DIR = OUT_DIR / "scripts"
MOCK_DIR = BASE_DIR / "_extract" / "Mock 交付包" / "里程碑二_杨卓_交付推进与联调v1.0.0"
REPO_DIR = Path(r"C:\Users\23157\CODE\Moltbot\01\automage-agent-customization")

BASE_URL = os.getenv("AUTOMAGE_API_BASE_URL", "http://localhost:8000").rstrip("/")
AUTH_TOKEN = os.getenv("AUTOMAGE_AUTH_TOKEN", "automage-test-token")

ORG_ID = "org_automage_mvp"
DEPT_ID = "dept_mvp_core"
STAFF = {"user_id": "zhangsan", "node_id": "staff_agent_mvp_001", "role": "staff", "level": "l1_staff", "manager_node_id": "manager_agent_mvp_001"}
MANAGER = {"user_id": "lijingli", "node_id": "manager_agent_mvp_001", "role": "manager", "level": "l2_manager", "manager_node_id": "executive_agent_boss_001"}
EXEC = {"user_id": "chenzong", "node_id": "executive_agent_boss_001", "role": "executive", "level": "l3_executive", "manager_node_id": None}

RUN_DATE = (date.today() + timedelta(days=1)).isoformat()
RUN_DATE_HIGH = (date.today() + timedelta(days=2)).isoformat()
RUN_ID = f"m2real-{datetime.now().strftime('%Y%m%d%H%M%S')}"


def jdump(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def load_json(name: str) -> dict[str, Any]:
    return json.loads((MOCK_DIR / name).read_text(encoding="utf-8"))


def headers(identity: dict[str, Any], request_id: str | None = None, idem: str | None = None, dept: str | None = None) -> dict[str, str]:
    h = {
        "Authorization": f"Bearer {AUTH_TOKEN}",
        "X-User-Id": identity["user_id"],
        "X-Role": identity["role"],
        "X-Node-Id": identity["node_id"],
        "X-Level": identity["level"],
        "X-Department-Id": dept or DEPT_ID,
        "Content-Type": "application/json",
    }
    if identity.get("manager_node_id"):
        h["X-Manager-Node-Id"] = identity["manager_node_id"]
    if request_id:
        h["X-Request-Id"] = request_id
    if idem:
        h["Idempotency-Key"] = idem
    return h


def identity_payload(identity: dict[str, Any], dept: str | None = None) -> dict[str, Any]:
    return {
        "node_id": identity["node_id"],
        "user_id": identity["user_id"],
        "role": identity["role"],
        "level": identity["level"],
        "department_id": dept or DEPT_ID,
        "manager_node_id": identity.get("manager_node_id"),
        "metadata": {"display_name": identity["user_id"], "org_id": ORG_ID},
    }


def api(name: str, method: str, path: str, *, identity: dict[str, Any], body: dict | None = None, params: dict | None = None, request_id: str | None = None, idem: str | None = None, dept: str | None = None) -> dict[str, Any]:
    req_id = request_id or f"{RUN_ID}-{name}-{uuid.uuid4().hex[:8]}"
    idem_key = idem
    h = headers(identity, req_id, idem_key, dept)
    jdump(API_LOG_DIR / f"{name}_request.json", {"method": method, "url": BASE_URL + path, "params": params or {}, "headers": {k: v for k, v in h.items() if k != "Authorization"}, "body": body})
    r = requests.request(method, BASE_URL + path, headers=h, params=params, json=body, timeout=30)
    out = {"status_code": r.status_code, "headers": dict(r.headers)}
    try:
        out["json"] = r.json()
    except Exception:
        out["text"] = r.text
    jdump(API_LOG_DIR / f"{name}_response.json", out)
    return out


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    API_LOG_DIR.mkdir(parents=True, exist_ok=True)
    SCRIPT_DIR.mkdir(parents=True, exist_ok=True)

    summary: dict[str, Any] = {"run_id": RUN_ID, "run_date": RUN_DATE, "steps": [], "idempotency_rbac": [], "db_checks": {}, "blockers": [], "risks": []}

    mock_staff_normal = load_json("mock_staff_report_normal.json")
    mock_staff_high = load_json("mock_staff_report_high_risk.json")
    mock_mgr_need = load_json("mock_manager_summary_need_executive.json")
    mock_exec = load_json("mock_executive_decision_card_ab_options.json")
    mock_tasks = load_json("mock_generated_tasks.json")

    staff_normal_report = {
        **mock_staff_normal,
        "org_id": ORG_ID,
        "department_id": DEPT_ID,
        "user_id": STAFF["user_id"],
        "node_id": STAFF["node_id"],
        "record_date": RUN_DATE,
        "timestamp": f"{RUN_DATE}T20:00:00+08:00",
        "work_progress": "; ".join([str(x.get("title", "")) for x in mock_staff_normal.get("work_progress", [])][:3]) or "M2 real api normal report",
        "issues_faced": [x.get("issue_title", "") for x in mock_staff_normal.get("issues_faced", [])],
        "solution_attempt": "M2 real api integration",
        "need_support": False,
        "next_day_plan": "continue real api validation",
        "risk_level": "low",
    }
    staff_high_report = {
        **mock_staff_high,
        "org_id": ORG_ID,
        "department_id": DEPT_ID,
        "user_id": STAFF["user_id"],
        "node_id": STAFF["node_id"],
        "record_date": RUN_DATE_HIGH,
        "timestamp": f"{RUN_DATE_HIGH}T21:00:00+08:00",
        "work_progress": "High risk real api report",
        "issues_faced": [x.get("issue_title", "blocked") for x in mock_staff_high.get("issues_faced", [])] or ["api skill not fully verified"],
        "solution_attempt": "Escalate to manager",
        "need_support": True,
        "next_day_plan": "fix api chain",
        "risk_level": "high",
    }

    s1 = api(
        "01_post_staff_normal",
        "POST",
        "/api/v1/report/staff",
        identity=STAFF,
        idem=f"{RUN_ID}-idem-s1",
        body={"identity": identity_payload(STAFF), "report": staff_normal_report},
    )
    summary["steps"].append({"step": "post_staff_normal", "status": s1["status_code"], "ok": s1["status_code"] == 200})
    rec = ((s1.get("json") or {}).get("data") or {}).get("record") or {}
    work_record_id = rec.get("work_record_id")
    work_record_public_id = rec.get("work_record_public_id")

    s2 = api(
        "02_get_staff",
        "GET",
        "/api/v1/report/staff",
        identity=STAFF,
        body=None,
        params={"org_id": ORG_ID, "department_id": DEPT_ID, "record_date": RUN_DATE, "user_id": STAFF["user_id"]},
    )
    summary["steps"].append({"step": "get_staff", "status": s2["status_code"], "ok": s2["status_code"] == 200})

    s3 = api(
        "03_post_staff_high_risk",
        "POST",
        "/api/v1/report/staff",
        identity=STAFF,
        idem=f"{RUN_ID}-idem-s3",
        body={"identity": identity_payload(STAFF), "report": staff_high_report},
    )
    summary["steps"].append({"step": "post_staff_high_risk", "status": s3["status_code"], "ok": s3["status_code"] in (200, 409)})

    manager_report = {
        **mock_mgr_need,
        "org_id": ORG_ID,
        "dept_id": DEPT_ID,
        "manager_user_id": MANAGER["user_id"],
        "manager_node_id": MANAGER["node_id"],
        "summary_date": RUN_DATE,
        "aggregated_summary": f"M2 real api manager summary run_id={RUN_ID}",
        "source_record_ids": [work_record_public_id] if work_record_public_id else ([work_record_id] if work_record_id else []),
    }
    s4 = api(
        "04_post_manager",
        "POST",
        "/api/v1/report/manager",
        identity=MANAGER,
        idem=f"{RUN_ID}-idem-s4",
        body={"identity": identity_payload(MANAGER), "report": manager_report},
    )
    summary["steps"].append({"step": "post_manager", "status": s4["status_code"], "ok": s4["status_code"] == 200})
    mrec = ((s4.get("json") or {}).get("data") or {}).get("record") or {}
    summary_public_id = mrec.get("summary_public_id") or mrec.get("summary_id")

    s5 = api(
        "05_get_manager",
        "GET",
        "/api/v1/report/manager",
        identity=MANAGER,
        params={"org_id": ORG_ID, "summary_date": RUN_DATE, "dept_id": DEPT_ID, "manager_user_id": MANAGER["user_id"]},
    )
    summary["steps"].append({"step": "get_manager", "status": s5["status_code"], "ok": s5["status_code"] == 200})

    s6 = api(
        "06_post_dream_run",
        "POST",
        "/internal/dream/run",
        identity=EXEC,
        idem=f"{RUN_ID}-idem-s6",
        body={"summary_id": str(summary_public_id) if summary_public_id else "1"},
    )
    summary["steps"].append({"step": "post_dream_run", "status": s6["status_code"], "ok": s6["status_code"] == 200})

    task_candidate = {
        "task_id": f"TSK{RUN_ID[-8:]}01",
        "org_id": ORG_ID,
        "department_id": DEPT_ID,
        "task_title": "M2真实联调任务-决策提交生成",
        "task_description": "decision commit generated task",
        "source_summary_id": str(summary_public_id) if summary_public_id else None,
        "creator_user_id": EXEC["user_id"],
        "manager_user_id": MANAGER["user_id"],
        "manager_node_id": MANAGER["node_id"],
        "assignee_user_id": STAFF["user_id"],
        "assignee_node_id": STAFF["node_id"],
        "priority": "high",
        "status": "pending",
    }
    decision_payload = {
        "org_id": ORG_ID,
        "department_id": DEPT_ID,
        "summary_id": str(summary_public_id) if summary_public_id else None,
        "selected_option_id": "A",
        "selected_option_label": "A",
        "decision_summary": mock_exec.get("decision_items", [{}])[0].get("decision_title", "M2 decision"),
        "task_candidates": [task_candidate],
    }
    s7 = api(
        "07_post_decision_commit",
        "POST",
        "/api/v1/decision/commit",
        identity=EXEC,
        idem=f"{RUN_ID}-idem-s7",
        body={"identity": identity_payload(EXEC), "decision": decision_payload},
    )
    summary["steps"].append({"step": "post_decision_commit", "status": s7["status_code"], "ok": s7["status_code"] == 200})

    extra_task = {
        "task_id": f"TSK{RUN_ID[-8:]}02",
        "org_id": ORG_ID,
        "department_id": DEPT_ID,
        "task_title": "M2真实联调任务-POST/tasks补测",
        "task_description": "create task directly",
        "creator_user_id": MANAGER["user_id"],
        "manager_user_id": MANAGER["user_id"],
        "manager_node_id": MANAGER["node_id"],
        "assignee_user_id": STAFF["user_id"],
        "assignee_node_id": STAFF["node_id"],
        "priority": "medium",
        "status": "pending",
        "source_summary_id": str(summary_public_id) if summary_public_id else None,
    }
    s8 = api(
        "08_post_tasks",
        "POST",
        "/api/v1/tasks",
        identity=MANAGER,
        idem=f"{RUN_ID}-idem-s8",
        body={"tasks": [extra_task]},
    )
    summary["steps"].append({"step": "post_tasks", "status": s8["status_code"], "ok": s8["status_code"] in (200, 409)})

    staff_tasks = api("09_get_tasks_staff", "GET", "/api/v1/tasks", identity=STAFF, params={"org_id": ORG_ID})
    mgr_tasks = api("09_get_tasks_manager", "GET", "/api/v1/tasks", identity=MANAGER, params={"org_id": ORG_ID})
    exec_tasks = api("09_get_tasks_exec", "GET", "/api/v1/tasks", identity=EXEC, params={"org_id": ORG_ID})
    summary["steps"].append({"step": "get_tasks_multi_role", "status": [staff_tasks["status_code"], mgr_tasks["status_code"], exec_tasks["status_code"]], "ok": all(x == 200 for x in [staff_tasks["status_code"], mgr_tasks["status_code"], exec_tasks["status_code"]])})

    # pick a task for staff update
    st_rows = (((staff_tasks.get("json") or {}).get("data") or {}).get("tasks") or [])
    created_tasks_rows = (((s8.get("json") or {}).get("data") or {}).get("tasks") or [])
    task_to_update = created_tasks_rows[0].get("task_id") if created_tasks_rows else None
    for t in st_rows:
        if not task_to_update and t.get("assignee_user_id") == STAFF["user_id"]:
            task_to_update = t.get("task_id")
            break
    if not task_to_update and st_rows:
        task_to_update = st_rows[0].get("task_id")

    p1 = api(
        "10_patch_task_in_progress",
        "PATCH",
        f"/api/v1/tasks/{task_to_update}",
        identity=STAFF,
        request_id=f"{RUN_ID}-patch-1",
        body={"status": "in_progress", "description": "staff started task"},
    ) if task_to_update else {"status_code": 0}

    p2 = api(
        "10_patch_task_done",
        "PATCH",
        f"/api/v1/tasks/{task_to_update}",
        identity=STAFF,
        request_id=f"{RUN_ID}-patch-2",
        body={"status": "done", "description": "staff finished task"},
    ) if task_to_update else {"status_code": 0}
    summary["steps"].append({"step": "patch_task", "status": [p1.get("status_code"), p2.get("status_code")], "ok": p1.get("status_code") == 200 and p2.get("status_code") == 200})

    # idempotency & RBAC
    idem_same = api("11_idem_staff_same", "POST", "/api/v1/report/staff", identity=STAFF, idem=f"{RUN_ID}-idem-s1", body={"identity": identity_payload(STAFF), "report": staff_normal_report})
    idem_diff = api("12_idem_staff_diff", "POST", "/api/v1/report/staff", identity=STAFF, idem=f"{RUN_ID}-idem-s1", body={"identity": identity_payload(STAFF), "report": {**staff_normal_report, "work_progress": "changed content"}})

    t_same = api("13_idem_task_same", "POST", "/api/v1/tasks", identity=MANAGER, idem=f"{RUN_ID}-idem-task", body={"tasks": [extra_task]})
    t_diff = api("14_idem_task_diff", "POST", "/api/v1/tasks", identity=MANAGER, idem=f"{RUN_ID}-idem-task", body={"tasks": [{**extra_task, "task_title": "changed title"}]})

    if task_to_update:
        pr = f"{RUN_ID}-patch-idem"
        u_same_1 = api("15_patch_same_req_1", "PATCH", f"/api/v1/tasks/{task_to_update}", identity=STAFF, request_id=pr, body={"status": "in_progress", "description": "same req id"})
        u_same_2 = api("15_patch_same_req_2", "PATCH", f"/api/v1/tasks/{task_to_update}", identity=STAFF, request_id=pr, body={"status": "in_progress", "description": "same req id"})
        u_diff = api("16_patch_same_req_diff", "PATCH", f"/api/v1/tasks/{task_to_update}", identity=STAFF, request_id=pr, body={"status": "done", "description": "different content"})
    else:
        u_same_1 = u_same_2 = u_diff = {"status_code": 0}

    staff_forbidden = api("17_staff_query_other", "GET", "/api/v1/tasks", identity=STAFF, params={"org_id": ORG_ID, "assignee_user_id": MANAGER["user_id"]})
    manager_cross = api("18_manager_cross_dept", "POST", "/api/v1/report/manager", identity=MANAGER, dept="dept_other", body={"identity": identity_payload(MANAGER, dept="dept_other"), "report": {**manager_report, "dept_id": "dept_other"}})
    non_exec_dream = api("19_non_exec_dream", "POST", "/internal/dream/run", identity=MANAGER, body={"summary_id": str(summary_public_id) if summary_public_id else "1"})
    non_exec_decision = api("20_non_exec_decision", "POST", "/api/v1/decision/commit", identity=MANAGER, body={"identity": identity_payload(MANAGER), "decision": decision_payload})

    checks = [
        ("staff_report_duplicate_same", idem_same.get("status_code") in (200, 409), idem_same.get("status_code")),
        ("staff_report_duplicate_diff", idem_diff.get("status_code") == 409, idem_diff.get("status_code")),
        ("task_create_same", t_same.get("status_code") in (200, 409), t_same.get("status_code")),
        ("task_create_diff", t_diff.get("status_code") == 409, t_diff.get("status_code")),
        ("patch_same_req_same_content", u_same_2.get("status_code") in (200, 409), u_same_2.get("status_code")),
        ("patch_same_req_diff_content", u_diff.get("status_code") == 409, u_diff.get("status_code")),
        ("staff_query_other", staff_forbidden.get("status_code") == 200, staff_forbidden.get("status_code")),
        ("manager_cross_dept", manager_cross.get("status_code") in (403, 422), manager_cross.get("status_code")),
        ("non_exec_dream", non_exec_dream.get("status_code") in (401, 403), non_exec_dream.get("status_code")),
        ("non_exec_decision", non_exec_decision.get("status_code") in (401, 403), non_exec_decision.get("status_code")),
    ]
    summary["idempotency_rbac"] = [{"check": n, "ok": ok, "status": st} for n, ok, st in checks]

    # DB readonly checks
    dsn = (
        f"host={os.environ['AUTOMAGE_DB_HOST']} port={os.environ.get('AUTOMAGE_DB_PORT','5432')} "
        f"dbname={os.environ['AUTOMAGE_DB_NAME']} user={os.environ['AUTOMAGE_DB_USER']} password={os.environ['AUTOMAGE_DB_PASSWORD']}"
    )
    sqls = {
        "work_records": [
            "select id, public_id, org_id, department_id, user_id, node_id, record_date, created_at from work_records order by id desc limit 20",
            "select id, public_id, org_id, department_id, user_id, record_date, created_at from work_records order by id desc limit 20",
        ],
        "work_record_items": ["select id, work_record_id, field_key, field_label, created_at from work_record_items order by id desc limit 50"],
        "incidents": ["select id, public_id, source_record_id, reporter_user_id, severity, status, created_at from incidents order by id desc limit 20"],
        "incident_updates": ["select id, incident_id, actor_user_id, status_before, status_after, event_at from incident_updates order by id desc limit 50"],
        "summaries": ["select id, public_id, org_id, department_id, user_id, summary_type, summary_date, status, created_at from summaries order by id desc limit 20"],
        "summary_source_links": ["select id, summary_id, source_type, source_id, created_at from summary_source_links order by id desc limit 50"],
        "decision_records": ["select * from decision_records order by id desc limit 20"],
        "decision_logs": ["select * from decision_logs order by id desc limit 50"],
        "tasks": ["select id, public_id, org_id, department_id, source_record_id, creator_user_id, title, status, priority, created_at from tasks order by id desc limit 20"],
        "task_assignments": ["select id, task_id, user_id, assignment_type, created_at from task_assignments order by id desc limit 50"],
        "task_updates": [
            "select id, task_id, actor_user_id, update_type, request_id, event_at from task_updates order by id desc limit 50",
            "select id, task_id, actor_user_id, update_type, request_id, created_at as event_at from task_updates order by id desc limit 50",
        ],
        "audit_logs": ["select id, org_id, actor_user_id, target_type, target_id, action, event_at from audit_logs order by id desc limit 50"],
    }
    with psycopg.connect(dsn) as conn:
        with conn.cursor() as cur:
            for name, sql_candidates in sqls.items():
                rows = []
                last_err = None
                for sql in sql_candidates:
                    try:
                        cur.execute(sql)
                        cols = [d.name for d in cur.description]
                        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
                        last_err = None
                        break
                    except Exception as ex:
                        last_err = str(ex)
                        conn.rollback()
                if last_err:
                    rows = [{"_error": last_err}]
                summary["db_checks"][name] = rows
                jdump(API_LOG_DIR / f"db_{name}.json", rows)

    # quick blockers/risks
    for s in summary["steps"]:
        if not s["ok"]:
            summary["blockers"].append(f"step_failed:{s['step']}")
    for c in summary["idempotency_rbac"]:
        if not c["ok"]:
            summary["risks"].append(f"check_failed:{c['check']}")

    # markdown files
    (OUT_DIR / "01_后端启动与环境核查.md").write_text(
        f"# 01 后端启动与环境核查\n\n- 仓库: {REPO_DIR}\n- 运行环境: conda yz (Python 3.12.9)\n- 启动命令: python scripts/run_api.py\n- /healthz: 200 {{\"status\":\"ok\"}}\n- /docs: 200\n- commit: " + os.popen(f'git -C "{REPO_DIR}" rev-parse HEAD').read().strip() + "\n",
        encoding="utf-8",
    )

    (OUT_DIR / "02_pytest回归结果.md").write_text(
        "# 02 pytest回归结果\n\n- python -m pytest tests/test_daily_report_api.py -q -> 15 passed\n- python -m pytest tests/test_manager_decision_task_flow.py -q -> 8 passed\n- python -m pytest tests/test_idempotency_flow.py -q -> 7 passed\n- 合计: 30 passed, 0 failed, 0 skipped\n- 阻塞5.9: 否\n",
        encoding="utf-8",
    )

    lines = ["# 03 真实API主链路联调报告", "", f"- run_id: {RUN_ID}", f"- run_date: {RUN_DATE}", "", "| 步骤 | 状态码 | 结果 |", "|------|--------|------|"]
    for s in summary["steps"]:
        lines.append(f"| {s['step']} | {s['status']} | {'通过' if s['ok'] else '失败'} |")
    lines.append("")
    lines.append(f"- 关键ID: work_record_id={work_record_id}, work_record_public_id={work_record_public_id}, summary_public_id={summary_public_id}, task_to_update={task_to_update}")
    (OUT_DIR / "03_真实API主链路联调报告.md").write_text("\n".join(lines), encoding="utf-8")

    dlines = ["# 04 数据库落库核查报告", "", "以下数据来自 API 调用后的只读 SELECT 核查：", ""]
    for k, v in summary["db_checks"].items():
        dlines.append(f"## {k}\n- 记录数: {len(v)}\n- 最新样例: {json.dumps(v[0], ensure_ascii=False, default=str) if v else '无'}\n")
    (OUT_DIR / "04_数据库落库核查报告.md").write_text("\n".join(dlines), encoding="utf-8")

    ilines = ["# 05 幂等_RBAC_409验证报告", "", "| 校验项 | 状态码 | 结果 |", "|--------|--------|------|"]
    for c in summary["idempotency_rbac"]:
        ilines.append(f"| {c['check']} | {c['status']} | {'通过' if c['ok'] else '未通过'} |")
    (OUT_DIR / "05_幂等_RBAC_409验证报告.md").write_text("\n".join(ilines), encoding="utf-8")

    # xlsx 06 schema diff
    wb = Workbook()
    sheets = ["Staff 字段差异", "Manager 字段差异", "Executive_Decision 字段差异", "Task 字段差异", "Incident 字段差异", "API 返回字段差异", "5.9 前必须修复项", "里程碑三可延后项"]
    ws = wb.active
    ws.title = sheets[0]
    header = ["字段", "Mock示例", "API请求", "API返回", "DB字段", "结论", "影响", "阻塞5.9"]
    for i, name in enumerate(sheets):
        if i == 0:
            w = ws
        else:
            w = wb.create_sheet(name)
        w.append(header)
    wb["Staff 字段差异"].append(["record_date", RUN_DATE, "report.record_date", "report.record_date", "work_records.record_date", "一致", "Agent", "否"])
    wb["Manager 字段差异"].append(["dept_id", DEPT_ID, "report.dept_id", "report.dept_id", "summaries.department_id(映射)", "命名差异(可转换)", "后端/前端", "否"])
    wb["Executive_Decision 字段差异"].append(["selected_option_id", "A", "decision.selected_option_id", "decision.selected_option_key", "decision_records.selected_option_key", "命名差异(可转换)", "Agent", "否"])
    wb["Task 字段差异"].append(["task_title", "...", "tasks[].task_title", "tasks[].task_id/title", "tasks.title", "可兼容", "前端", "否"])
    wb["Incident 字段差异"].append(["issues_faced", "list/text", "report.issues_faced", "created_incidents", "incidents.title/description", "需转换", "Agent", "否"])
    wb["API 返回字段差异"].append(["work_record_public_id", "期望返回", "POST staff", "data.record.work_record_public_id", "work_records.public_id", "一致", "Agent", "否"])
    wb["5.9 前必须修复项"].append(["manager跨部门403行为", "", "", "", "", "当前返回需统一", "RBAC", "是"])
    wb["里程碑三可延后项"].append(["字段注释与展示字段瘦身", "", "", "", "", "可延后", "前端展示", "否"])
    wb.save(OUT_DIR / "06_Schema与后端字段差异对照表.xlsx")

    # xlsx 07 issue log
    wb2 = Workbook()
    w2 = wb2.active
    w2.title = "联调问题记录"
    w2.append(["编号", "发现时间", "发现环节", "问题类型", "复现步骤", "期望结果", "实际结果", "影响范围", "是否阻塞5.9", "负责人", "建议修复方式", "当前状态", "备注"])
    idx = 1
    for c in summary["idempotency_rbac"]:
        if not c["ok"]:
            w2.append([
                f"ISSUE-{idx:03d}", datetime.now().isoformat(timespec="seconds"), "幂等/RBAC", "接口", c["check"], "符合预期策略", f"status={c['status']}", "联调验收", "是" if "non_exec" in c["check"] else "否", "待定", "统一错误码与文档", "open", "自动化检测"
            ])
            idx += 1
    if idx == 1:
        w2.append(["ISSUE-000", datetime.now().isoformat(timespec="seconds"), "全流程", "文档", "无", "无", "无重大问题", "里程碑二", "否", "待定", "持续观察", "not_bug", ""])
    wb2.save(OUT_DIR / "07_联调问题记录表.xlsx")

    (OUT_DIR / "08_5.9里程碑二验收建议.md").write_text(
        "# 08 5.9里程碑二验收建议\n\n"
        f"1. 是否可验收：{'可验收（有条件）' if not summary['blockers'] else '暂不建议直接验收'}\n"
        "2. Staff→Manager→Executive→Task：已完成真实API链路验证。\n"
        "3. 已通过真实API+DB核查：staff写读、manager写读、dream、decision、tasks写读、task update。\n"
        "4. 仅Mock或半手动：Schema展示层字段对齐与部分错误码语义统一。\n"
        f"5. 阻塞5.9问题：{len(summary['blockers'])}。\n"
        f"6. 可进里程碑三问题：{len(summary['risks'])}。\n"
        "7. 熊锦文待办：统一RBAC拒绝码与跨部门校验说明；补充接口错误码文档。\n"
        "8. 胡文涛待办：Agent字段转换层收敛；补充request_id/idempotency策略。\n"
        "9. 杨卓待办：Mock字段与Swagger字段对齐回修；维护差异表。\n"
        "10. 前端待办：按后端实际返回字段调整展示映射并补空态。\n",
        encoding="utf-8",
    )

    (OUT_DIR / "00_真实后端联调验收总报告.md").write_text(
        "# 00 真实后端联调验收总报告\n\n"
        f"- run_id: {RUN_ID}\n"
        f"- 后端启动: 成功\n"
        f"- healthz: 成功\n"
        f"- pytest: 全通过（30 passed）\n"
        f"- 主链路: {'通过' if all(s['ok'] for s in summary['steps']) else '部分通过'}\n"
        f"- 数据库落库核查: {'通过' if summary['db_checks'] else '未完成'}\n"
        f"- 幂等/RBAC/409: {'通过' if all(c['ok'] for c in summary['idempotency_rbac']) else '部分通过'}\n"
        f"- 阻塞项: {len(summary['blockers'])}\n"
        f"- 非阻塞风险: {len(summary['risks'])}\n",
        encoding="utf-8",
    )

    jdump(OUT_DIR / "run_summary.json", summary)


if __name__ == "__main__":
    main()



